import openpyxl


class HomePageData:
   @staticmethod
   def get_test_data():
      homepageData = []
      dataDict = {}
      dataFile = openpyxl.load_workbook('C:\\Users\\Babis\\PycharmProjects'
                                        '\\pythonSeleniumComplete\\testData\\Data.xlsx')
      dataSheet = dataFile.active
      for rowIndex in range(1, dataSheet.max_row + 1):
         if 'TestCase' in dataSheet.cell(row=rowIndex, column=1).value:
            for columnIndex in range(2, dataSheet.max_column + 1):
               dataDict[dataSheet.cell(row=1, column=columnIndex).value] =\
                  dataSheet.cell(row=rowIndex, column=columnIndex).value
            homepageData.append(dataDict)
      return homepageData
